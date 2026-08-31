# -*- coding: utf-8 -*-
"""WPS/Excel 工作表回写：按机子号匹配行，写入「登入情况」列。"""

from __future__ import annotations

import threading
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pythoncom
    import win32com.client

    HAS_COM = True
except ImportError:
    HAS_COM = False

_excel_lock = threading.Lock()

DEFAULT_MACHINE_HEADERS = [
    "机子号",
    "环境名称",
    "环境",
    "containerName",
    "machine_id",
    "Machine",
    "HubStudio",
]
DEFAULT_STATUS_HEADERS = ["登入情况", "登录情况"]

# WPS 表格 / Microsoft Excel COM ProgID
_SPREADSHEET_PROGIDS = (
    "Ket.Application",   # WPS 表格（常见）
    "ET.Application",    # 金山 ET / 旧版 WPS
    "kwps.Application",  # WPS Office
    "Excel.Application", # Microsoft Excel
)


def normalize_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_machine_id(value: Any) -> str:
    return normalize_cell_text(value)


def normalize_path(path: Path | str) -> str:
    try:
        return str(Path(path).resolve()).lower()
    except Exception:
        return str(path).lower()


NOT_FOUND_PREFIX = "表中未找到机子号"


def cell_matches_machine(cell_value: Any, machine_id: str) -> bool:
    """机子号精确匹配（忽略大小写与首尾空格）。"""
    cell_text = normalize_machine_id(cell_value)
    target = normalize_machine_id(machine_id)
    if not cell_text or not target:
        return False
    return cell_text.lower() == target.lower()


def header_matches(val: str, names: list[str]) -> bool:
    if not val:
        return False
    if val in names:
        return True
    for name in names:
        if name in val:
            return True
    return False


def detect_header_row_from_matrix(
    get_cell,
    max_row: int,
    max_col: int,
    max_scan_rows: int = 10,
) -> int:
    limit = min(max_scan_rows, max_row)
    for row in range(1, limit + 1):
        for col in range(1, max_col + 1):
            val = normalize_cell_text(get_cell(row, col))
            if not val:
                continue
            if header_matches(val, DEFAULT_MACHINE_HEADERS) or header_matches(
                val, DEFAULT_STATUS_HEADERS
            ):
                return row
            if "机子" in val or "登入情况" in val or "登录情况" in val:
                return row
    return 1


def find_column_from_matrix(
    get_cell,
    header_row: int,
    max_col: int,
    names: list[str],
) -> int | None:
    for col in range(1, max_col + 1):
        val = normalize_cell_text(get_cell(header_row, col))
        if header_matches(val, names):
            return col
    return None


def find_machine_row_from_matrix(
    get_cell,
    machine_id: str,
    machine_col: int | None,
    header_row: int,
    max_row: int,
    max_col: int,
) -> int | None:
    """先在机子号列查找；未命中则扫描整行所有单元格。"""
    if machine_col:
        for row in range(header_row + 1, max_row + 1):
            if cell_matches_machine(get_cell(row, machine_col), machine_id):
                return row
    for row in range(header_row + 1, max_row + 1):
        for col in range(1, max_col + 1):
            if cell_matches_machine(get_cell(row, col), machine_id):
                return row
    return None


def resolve_sheet_layout(
    get_cell,
    max_row: int,
    max_col: int,
    machine_headers: list[str] | None = None,
    status_headers: list[str] | None = None,
) -> dict[str, int]:
    """解析表头：机子号列 + 登入情况列（同一行）。"""
    header_row = detect_header_row_from_matrix(get_cell, max_row, max_col)
    machine_col = find_column_from_matrix(
        get_cell, header_row, max_col, machine_headers or DEFAULT_MACHINE_HEADERS
    )
    status_col = find_column_from_matrix(
        get_cell, header_row, max_col, status_headers or DEFAULT_STATUS_HEADERS
    )
    if status_col is None:
        raise ValueError("未找到登入情况列（表头需含：登入情况/登录情况）")
    return {
        "header_row": header_row,
        "machine_col": machine_col,
        "status_col": status_col,
    }


def write_status_for_machine(
    get_cell,
    set_cell,
    machine_id: str,
    value: str,
    max_row: int,
    max_col: int,
    *,
    machine_headers: list[str] | None = None,
    status_headers: list[str] | None = None,
) -> tuple[bool, str]:
    """查号完成后：定位机子号所在行 → 写入同行「登入情况」单元格。"""
    layout = resolve_sheet_layout(
        get_cell,
        max_row,
        max_col,
        machine_headers=machine_headers,
        status_headers=status_headers,
    )
    row = find_machine_row_from_matrix(
        get_cell,
        machine_id,
        layout["machine_col"],
        layout["header_row"],
        max_row,
        max_col,
    )
    if row is None:
        return False, f"{NOT_FOUND_PREFIX}: {machine_id}"

    set_cell(row, layout["status_col"], value)
    col_letter = layout["status_col"]
    return True, f"已定位机子号第 {row} 行，写入登入情况列第 {col_letter} 列"


# ---------- openpyxl ----------


def find_header_row(ws, max_scan_rows: int = 10) -> int:
    limit = min(max_scan_rows, ws.max_row or 1)

    def get_cell(r, c):
        return ws.cell(r, c).value

    return detect_header_row_from_matrix(
        get_cell, limit, ws.max_column or 1, max_scan_rows
    )


def find_column(ws, header_row: int, names: list[str]) -> int | None:
    max_col = ws.max_column or 1

    def get_cell(r, c):
        return ws.cell(r, c).value

    return find_column_from_matrix(get_cell, header_row, max_col, names)


def resolve_sheet_column_map(
    ws,
    machine_headers: list[str] | None = None,
    status_headers: list[str] | None = None,
) -> dict[str, int]:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    def get_cell(r, c):
        return ws.cell(r, c).value

    layout = resolve_sheet_layout(
        get_cell, max_row, max_col, machine_headers, status_headers
    )
    if layout["machine_col"] is None:
        pass  # 允许无表头机子号列，后续全表扫描
    return layout


def find_machine_row(
    ws,
    machine_id: str,
    machine_col: int | None,
    header_row: int,
) -> int | None:
    max_row = ws.max_row or header_row
    max_col = ws.max_column or 1

    def get_cell(r, c):
        return ws.cell(r, c).value

    return find_machine_row_from_matrix(
        get_cell, machine_id, machine_col, header_row, max_row, max_col
    )


# ---------- COM (WPS / Excel 已打开) ----------


def _get_spreadsheet_app():
    if not HAS_COM:
        return None, None
    last_exc: Exception | None = None
    for progid in _SPREADSHEET_PROGIDS:
        try:
            app = win32com.client.GetActiveObject(progid)
            return app, progid
        except Exception as exc:
            last_exc = exc
            try:
                app = win32com.client.Dispatch(progid)
                if getattr(app, "Workbooks", None) and app.Workbooks.Count >= 0:
                    return app, progid
            except Exception as exc2:
                last_exc = exc2
    return None, last_exc


def _find_open_workbook(app, file_path: Path):
    target = normalize_path(file_path)
    target_name = Path(file_path).name.lower()
    try:
        count = int(app.Workbooks.Count)
    except Exception:
        return None
    fallback = None
    for idx in range(1, count + 1):
        try:
            wb = app.Workbooks(idx)
            full = normalize_path(wb.FullName)
            if full == target:
                return wb
            if Path(wb.FullName).name.lower() == target_name:
                fallback = wb
        except Exception:
            continue
    return fallback


def _com_sheet_bounds(ws) -> tuple[int, int]:
    try:
        used = ws.UsedRange
        max_row = int(used.Row + used.Rows.Count - 1)
        max_col = int(used.Column + used.Columns.Count - 1)
        return max(max_row, 1), max(max_col, 1)
    except Exception:
        return 5000, 50


def _update_via_com(
    path: Path,
    sheet_name: str,
    machine_id: str,
    value: str,
    *,
    machine_headers: list[str] | None,
    status_headers: list[str] | None,
) -> tuple[bool, str]:
    if not HAS_COM:
        return False, "未安装 pywin32，无法在 WPS 打开时写入"

    pythoncom.CoInitialize()
    try:
        app, progid_or_err = _get_spreadsheet_app()
        if app is None:
            return False, f"未检测到 WPS/Excel 进程: {progid_or_err}"

        wb = _find_open_workbook(app, path)
        if wb is None:
            return False, "WPS/Excel 中未打开目标文件，请先在 WPS 中打开该表格"

        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            return False, f"工作表不存在: {sheet_name}"

        max_row, max_col = _com_sheet_bounds(ws)

        def get_cell(r, c):
            return ws.Cells(r, c).Value

        def set_cell(r, c, val):
            ws.Cells(r, c).Value = val

        ok, detail = write_status_for_machine(
            get_cell,
            set_cell,
            machine_id,
            value,
            max_row,
            max_col,
            machine_headers=machine_headers,
            status_headers=status_headers,
        )
        if not ok:
            return False, detail

        try:
            wb.Save()
        except Exception as exc:
            return False, f"COM 保存失败: {exc}"
        return True, f"{detail}（WPS 已打开）"
    finally:
        pythoncom.CoUninitialize()


def _list_sheets_via_com(path: Path) -> tuple[list[str], str | None]:
    if not HAS_COM:
        return [], None
    pythoncom.CoInitialize()
    try:
        app, _ = _get_spreadsheet_app()
        if app is None:
            return [], None
        wb = _find_open_workbook(app, path)
        if wb is None:
            return [], None
        names: list[str] = []
        try:
            count = int(wb.Worksheets.Count)
            for idx in range(1, count + 1):
                names.append(str(wb.Worksheets(idx).Name))
        except Exception:
            pass
        return names, None
    finally:
        pythoncom.CoUninitialize()


# ---------- public API ----------


def list_sheet_names(file_path: Path | str) -> tuple[list[str], str | None]:
    path = Path(file_path)
    if not path.is_file():
        return [], f"Excel 文件不存在: {path}"

    com_names, _ = _list_sheets_via_com(path)
    if com_names:
        return com_names, None

    if not HAS_OPENPYXL:
        return [], "未安装 openpyxl"
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names, None
    except PermissionError:
        com_names, _ = _list_sheets_via_com(path)
        if com_names:
            return com_names, None
        return [], "Excel 文件被占用；请先在 WPS 中打开该文件，或关闭后重试"
    except Exception as exc:
        return [], f"无法读取 Excel: {exc}"


def build_status_value(status: str, detail: str = "", *, include_detail: bool = False) -> str:
    status = normalize_cell_text(status)
    detail = normalize_cell_text(detail)
    if include_detail and detail:
        return f"{status} | {detail}"
    return status


def update_machine_status(
    file_path: Path | str,
    sheet_name: str,
    machine_id: str,
    status: str,
    detail: str = "",
    *,
    machine_headers: list[str] | None = None,
    status_headers: list[str] | None = None,
    include_detail: bool = False,
    write_while_open: bool = True,
) -> tuple[bool, str]:
    """按机子号匹配行，写入登入情况列。优先写入 WPS/Excel 已打开的工作簿。"""
    path = Path(file_path)
    if not path.is_file():
        return False, f"Excel 文件不存在: {path}"

    sheet_name = normalize_cell_text(sheet_name)
    if not sheet_name:
        return False, "未指定工作表名"

    value = build_status_value(status, detail, include_detail=include_detail)

    with _excel_lock:
        if write_while_open:
            ok, msg = _update_via_com(
                path,
                sheet_name,
                machine_id,
                value,
                machine_headers=machine_headers,
                status_headers=status_headers,
            )
            if ok:
                return True, msg

        if not HAS_OPENPYXL:
            if write_while_open:
                return False, msg
            return False, "未安装 openpyxl"

        try:
            wb = load_workbook(path)
        except PermissionError:
            if write_while_open:
                ok, com_msg = _update_via_com(
                    path,
                    sheet_name,
                    machine_id,
                    value,
                    machine_headers=machine_headers,
                    status_headers=status_headers,
                )
                if ok:
                    return True, com_msg
                return False, com_msg or "Excel 被占用且 COM 写入失败，请确认 WPS 已打开该文件"
            return False, "Excel 文件被占用，请关闭 WPS 后重试"
        except Exception as exc:
            return False, f"无法打开 Excel: {exc}"

        if sheet_name not in wb.sheetnames:
            wb.close()
            return False, f"工作表不存在: {sheet_name}"

        ws = wb[sheet_name]
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        def get_cell(r, c):
            return ws.cell(r, c).value

        def set_cell(r, c, val):
            ws.cell(r, c).value = val

        try:
            ok, detail = write_status_for_machine(
                get_cell,
                set_cell,
                machine_id,
                value,
                max_row,
                max_col,
                machine_headers=machine_headers,
                status_headers=status_headers,
            )
        except ValueError as exc:
            wb.close()
            return False, str(exc)
        if not ok:
            wb.close()
            return False, detail

        try:
            wb.save(path)
        except PermissionError:
            wb.close()
            if write_while_open:
                ok, com_msg = _update_via_com(
                    path,
                    sheet_name,
                    machine_id,
                    value,
                    machine_headers=machine_headers,
                    status_headers=status_headers,
                )
                if ok:
                    return True, com_msg
            return False, "保存失败，请确认 WPS 已打开该文件"
        except Exception as exc:
            wb.close()
            return False, f"保存失败: {exc}"
        wb.close()
        return True, detail


def is_machine_not_found_message(msg: str) -> bool:
    return normalize_cell_text(msg).startswith(NOT_FOUND_PREFIX)


def excel_sync_config(config: dict) -> dict[str, Any]:
    raw = config.get("excel_sync") or {}
    return {
        "enabled": bool(raw.get("enabled", False)),
        "file_path": normalize_cell_text(raw.get("file_path")),
        "machine_headers": raw.get("machine_headers") or DEFAULT_MACHINE_HEADERS,
        "status_headers": raw.get("status_headers") or DEFAULT_STATUS_HEADERS,
        "include_detail": bool(raw.get("include_detail", False)),
        "write_while_open": bool(raw.get("write_while_open", True)),
    }
